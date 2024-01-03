import serial
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime as dt

MAX_COUNTER = int(input("How many measurements?: "))

serial_connection = serial.Serial("COM5")
filename = r"D:\cansat\tests\Tests.xlsx"
wb = load_workbook(filename)

time = f"{dt.now().day}.{dt.now().month}.{dt.now().year}-{dt.now().hour};{dt.now().minute};{dt.now().second}"

wb.create_sheet(time)
ws = wb[time]

ws["A1"].value = "Temperature"
ws["B1"].value = "Pressure"

counter = 1
row = 2
while counter <= MAX_COUNTER:
    data = serial_connection.readline()
    print(data)
    data = str(data)[2:]
    data = data[:-5]
    data = data.split(", ")
    
    for i in range(2):
        col = get_column_letter(i+1)
        ws[col+str(row)].value = data[i]

    if data == b"EOF":
        break
    counter += 1
    row += 1

wb.save(filename)
serial_connection.close()